import requests
from openpyxl import Workbook
from openpyxl import load_workbook
import json
import sys
import shutil


shutil.copy("Sample.xlsx","Test_Report.xlsx")
wb = load_workbook(filename = 'Test_Report.xlsx')
ws = wb["Test Data"]
url = "http://10.197.55.80:8080/v1/vcp/action?command="


for x in range(2, ws.max_row+1):
    finalUrl = url + unicode(ws.cell(row=x,column=2).value) + " " + unicode(ws.cell(row=x,column=3).value)
    print finalUrl
    
    try:
        r = requests.get(finalUrl)
        r.raise_for_status()
        data = r.text
        print data
        ws.cell(row=x,column=5).value = data
        if(ws.cell(row=x,column=4).value == ws.cell(row=x,column=5).value):
            ws.cell(row=x,column=6).value = 'PASS'
        else:
            ws.cell(row=x,column=6).value = 'FAIL'
            ws.cell(row=x,column=7).value = 'Response Assertion Failure'
        print "============================="

    except requests.exceptions.HTTPError as errh:
        print ("Http Error:",errh)
        ws.cell(row=x,column=6).value = 'FAIL'
        ws.cell(row=x,column=7).value = 'Http Error:' + str(errh)
        print "============================="
    except requests.exceptions.ConnectionError as errc:
        print ("Error Connecting:",errc)
        ws.cell(row=x,column=6).value = 'FAIL'
        ws.cell(row=x,column=7).value = 'Error Connecting:' + str(errc)
        print "============================="
    except requests.exceptions.Timeout as errt:
        print ("Timeout Error:",errt)
        ws.cell(row=x,column=6).value = 'FAIL'
        ws.cell(row=x,column=7).value = 'Timeout Error:' + str(errt)
        print "============================="

# Save the file
wb.save("Test_Report.xlsx")
